import urllib.request
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from fake_useragent import UserAgent as UA
import time
from googletrans import Translator

#Обход защиты
UA = UA.random

#настройки окна
options = Options()
options.add_argument("--headless")
options.add_argument("--disable-gpu")  # Отключаем GPU для стабильности
options.add_argument("--window-size=1920,1080")  # Размер окна
options.add_argument(f"user-agent={UA}")

#пусть до драйвера хром, создание игрока
driver = webdriver.Chrome( options=options)

#переменные
row_count = 313
translator = Translator()

def card(elements):
    for element in elements:

        #Открытие рабочей таблицы 
        workbook = openpyxl.load_workbook("C:\python\JohnDeere\JohnDeere.xlsx") #Путь к таблице с перенесённым товаром 
        ws = workbook.active

        href = element.get_attribute("href")

        #Основной Скрипт

        driver.get(href)
        time.sleep(3)

        try:
            title = driver.find_element(By.XPATH, '//h1[@data-testid="productName"]')
            article_number = title.text.split(": ")
        except:
            print(f"Ошибка в товаре {href}")
            continue
        
        try:
            photo = driver.find_element(By.XPATH, f'//img[contains(@alt, "{article_number[0]}")]')
            urllib.request.urlretrieve(photo.get_attribute("src"), f"C:\python\JohnDeere\photo\{article_number[0]}.jpg")
        except:
            urllib.request.urlretrieve(("https://optim.tildacdn.com/stor6466-6633-4435-a665-383462376632/-/format/webp/98648328.jpg"), f"C:\python\JohnDeere\photo\{article_number[0]}.jpg")

        tanslate_title = translator.translate(title.text.split(": ")[1], src='en', dest='ru')
        title = tanslate_title.text + " " + article_number[0]

        global row_count
        ws.cell(row=row_count, column=1, value=title)
        ws.cell(row=row_count, column=2, value=article_number[0])

        row_count += 1
        
        #Сохранеиние рабочей таблицы 
        workbook.save("C:\python\JohnDeere\JohnDeere.xlsx")
